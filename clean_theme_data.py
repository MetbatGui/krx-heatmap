import pandas as pd
import os
import unicodedata

def normalize_str(s):
    if pd.isna(s) or str(s).lower() == 'nan':
        return None
    val = str(s).strip()
    if not val:
        return None
    return unicodedata.normalize('NFC', val)

def clean_theme_data():
    input_path = 'data/theme_data/unique_theme_heatmap_data.xlsx'
    output_path = 'data/theme_data/unique_theme_heatmap_data.xlsx'

    print(f"Loading data from {input_path}...")
    try:
        xl = pd.ExcelFile(input_path)
        df_detail = pd.read_excel(xl, '테마상세')
        df_duplicates = pd.read_excel(xl, '중복종목')
    except Exception as e:
        print(f"Error loading excel file: {e}")
        return

    # 1. Convert '테마상세' to a dictionary of sets for efficient lookup and modification
    # Structure: {Theme: {Stock1, Stock2, ...}}
    print("Processing detailed theme data...")
    theme_dict = {}
    valid_themes = set()
    
    for col in df_detail.columns:
        normalized_col = normalize_str(col)
        valid_themes.add(normalized_col)
        
        # Get valid stock names (drop NaNs)
        stocks = set([normalize_str(x) for x in df_detail[col].dropna().tolist()])
        theme_dict[normalized_col] = stocks

    # 2. Process duplicates
    print("Applying duplicates filter...")
    
    # Trackers for reporting
    invalid_theme_report = []

    # Identify duplicate/theme mapping columns (All columns except '종목명')
    theme_cols = [c for c in df_duplicates.columns if c != '종목명']
    print(f"Theme columns detected: {len(theme_cols)} columns")
    
    processed_count = 0
    
    for idx, row in df_duplicates.iterrows():
        stock_name = normalize_str(row['종목명'])
        
        # Find the target theme for this stock
        target_theme = None
        for col in theme_cols:
            candidate_theme = normalize_str(row[col])
            
            if not candidate_theme:
                continue

            if candidate_theme in valid_themes:
                target_theme = candidate_theme
                break
            else:
                # Warning: Theme specified but not found in valid themes
                invalid_theme_report.append({
                    '종목명': stock_name,
                    '입력테마': candidate_theme,
                    '입력컬럼': col
                })

        if target_theme:
            processed_count += 1
            # Action: Ensure stock is in target_theme, remove from all others
            
            # 1. Remove from ALL themes first (to ensure no duplicates remain)
            for theme in theme_dict:
                if stock_name in theme_dict[theme]:
                    theme_dict[theme].remove(stock_name)
            
            # 2. Add to target theme
            theme_dict[target_theme].add(stock_name)
        
        if stock_name == 'LG전자':
             # Keep debug just in case
             pass

    print(f"Processed {processed_count} stocks from duplicates list.")

    # 3. Analyze Unresolved Duplicates
    print("Analyzing unresolved duplicates...")
    stock_to_themes = {}
    for theme, stocks in theme_dict.items():
        for stock in stocks:
            if stock not in stock_to_themes:
                stock_to_themes[stock] = []
            stock_to_themes[stock].append(theme)
            
    unresolved_duplicates_report = []
    for stock, themes in stock_to_themes.items():
        if len(themes) > 1:
            unresolved_duplicates_report.append({
                '종목명': stock,
                '소속테마수': len(themes),
                '소속테마목록': ', '.join(themes)
            })

    # 4. Reconstruct DataFrame
    print("Reconstructing DataFrame...")
    # Convert sets back to lists and sort them (optional, but good for consistency)
    result_data = {theme: sorted(list(stocks)) for theme, stocks in theme_dict.items()}
    
    # Create DataFrame (will automatically pad with NaNs)
    df_result = pd.DataFrame(dict([ (k,pd.Series(v)) for k,v in result_data.items() ]))

    # 5. Save result
    print(f"Saving result to {output_path}...")
    try:
        # Save validation: Ensure columns match original order
        df_result = df_result.reindex(columns=df_detail.columns)
        
        # Create DataFrames for reports
        df_invalid = pd.DataFrame(invalid_theme_report)
        df_unresolved = pd.DataFrame(unresolved_duplicates_report)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_result.to_excel(writer, sheet_name='테마상세', index=False)
            df_duplicates.to_excel(writer, sheet_name='중복종목', index=False)
            
            # Save new reports
            if not df_unresolved.empty:
                df_unresolved.to_excel(writer, sheet_name='중복_미제거', index=False)
            else:
                 pd.DataFrame({'Info': ['모든 중복이 해결되었습니다.']}).to_excel(writer, sheet_name='중복_미제거', index=False)
                 
            if not df_invalid.empty:
                df_invalid.to_excel(writer, sheet_name='테마_오표기', index=False)
            else:
                 pd.DataFrame({'Info': ['테마 오표기가 없습니다.']}).to_excel(writer, sheet_name='테마_오표기', index=False)
                 
        # 6. Apply Highlighting (Red for duplicates)
        if not df_unresolved.empty:
            print("Applying red highlights to duplicates...")
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            
            wb = load_workbook(output_path)
            ws = wb['테마상세']
            
            red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            
            # Get list of duplicate stocks
            dup_stocks = set(df_unresolved['종목명'].apply(normalize_str))
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_val = normalize_str(cell.value)
                        if cell_val in dup_stocks:
                            cell.fill = red_fill
            
            wb.save(output_path)
            print("Highlights applied.")
            
        print("Done.")
    except Exception as e:
        print(f"Error saving file: {e}")

if __name__ == "__main__":
    clean_theme_data()
