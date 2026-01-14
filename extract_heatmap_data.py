import os
import glob
import pandas as pd
from bs4 import BeautifulSoup
from collections import defaultdict
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def extract_stocks_from_html(file_path):
    """
    HTML 파일에서 종목명과 티커를 추출하여 DataFrame으로 반환합니다.
    컬럼명은 파일명(확장자 제외)입니다.
    """
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return None

    file_name = os.path.splitext(os.path.basename(file_path))[0]
    
    with open(file_path, 'r', encoding='utf-8') as f:
        html_content = f.read()

    soup = BeautifulSoup(html_content, 'html.parser')
    rows = soup.select('tr.stockTrMobile')

    stock_data = []

    for row in rows:
        p_tags = row.select('td p.stockInfoMobile')
        if len(p_tags) >= 2:
            stock_name = p_tags[0].text.strip()
            stock_data.append(stock_name)

    # DataFrame 생성
    df = pd.DataFrame({file_name: stock_data})
    # 중복 제거
    df = df.drop_duplicates()
    
    return df

if __name__ == "__main__":
    # Add src to python path to import config
    sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))
    try:
        from theme_config import PRIORITY_THEMES, THEME_RENAME
    except ImportError:
        print("Warning: Could not import theme_config. Using defaults.")
        PRIORITY_THEMES = []
        THEME_RENAME = {}

    # 타겟 파일 리스트 (glob 사용)
    target_files = glob.glob(r'data/theme_html/*.html')
    
    print(f"Loading {len(target_files)} theme files...")
    
    # 테마별 종목 리스트 수집
    theme_data_list = [] # [{'theme': name, 'stocks': [list]}, ...]
    
    for file_path in target_files:
        df = extract_stocks_from_html(file_path)
        if df is not None:
            original_theme_name = df.columns[0]
            # 테마명 변경 적용
            theme_name = THEME_RENAME.get(original_theme_name, original_theme_name)
            
            stocks = df[original_theme_name].tolist() # df column name is still original
            if stocks:
                theme_data_list.append({'theme': theme_name, 'stocks': stocks})
    
    # 정렬 및 우선순위 처리
    # 1. 우선순위 테마 추출
    priority_list = []
    normal_list = []
    
    for item in theme_data_list:
        if item['theme'] in PRIORITY_THEMES:
            priority_list.append(item)
        else:
            normal_list.append(item)
            
    # 우선순위 리스트는 Config 순서에 맞게 정렬
    priority_list.sort(key=lambda x: PRIORITY_THEMES.index(x['theme']))
    
    # 2. 일반 리스트는 종목 수 많은 순으로 정렬
    normal_list.sort(key=lambda x: len(x['stocks']), reverse=True)
    
    # 병합 (우선순위 먼저)
    sorted_theme_list = priority_list + normal_list
    
    # 3. 앞에서부터 중복 제거하며 수집 (Greedy Selection)
    final_rows = []
    seen_stocks = set()
    total_count = 0
    selected_theme_count = 0
    
    print("\nProcessing Greedy Selection...")
    
    for item in sorted_theme_list:
        theme = item['theme']
        original_stocks = item['stocks']
        
        # 이미 등록된 종목 제외
        new_stocks = [s for s in original_stocks if s not in seen_stocks]
        
        if not new_stocks:
            continue
            
        # 테마당 최대 33개로 제한 (우선순위 테마는 제한을 어떻게 할지? 일단 동일하게 적용)
        if len(new_stocks) > 33:
            new_stocks = new_stocks[:33]
            
        # 결과에 추가
        for stock in new_stocks:
            final_rows.append({'테마': theme, '종목명': stock})
            seen_stocks.add(stock)
            
        selected_theme_count += 1
        total_count = len(seen_stocks)
        
        # 4. 200개가 넘을 때까지
        if total_count > 200:
            print(f"Reached target count: {total_count} stocks from {selected_theme_count} themes.")
            break
            
    if total_count <= 200:
        print(f"Warning: Only found {total_count} stocks total (Target > 200).")

    # DataFrame 생성
    df_result = pd.DataFrame(final_rows)
    
    # 엑셀 저장
    output_file = 'heatmap_data.xlsx'
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
        except PermissionError:
            print(f"Error: {output_file} is open. Please close it.")
            sys.exit(1)
            
    try:
        # 데이터 저장
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_result.to_excel(writer, sheet_name='테마와 종목명', index=False)
            
        # 컬럼 너비 조정
        wb = load_workbook(output_file)
        ws = wb['테마와 종목명']
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.5
            ws.column_dimensions[column].width = adjusted_width
            
        wb.save(output_file)
        print(f"\nSaved to {output_file}")
        print(f"Total Themes Used: {selected_theme_count}")
        print(f"Total Stocks Selected: {total_count}")
        
    except Exception as e:
        print(f"Error saving excel: {e}")
