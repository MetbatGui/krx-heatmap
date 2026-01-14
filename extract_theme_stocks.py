import os
import pandas as pd
from bs4 import BeautifulSoup

def extract_stocks_from_html(file_path):
    """
    HTML 파일에서 종목명과 티커를 추출하여 DataFrame으로 반환합니다.
    컬럼명은 파일명(확장자 제외)입니다.
    """
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return None

    file_name = os.path.splitext(os.path.basename(file_path))[0]
    
    with open(file_path, 'r', encoding='utf-8') as f:
        html_content = f.read()

    soup = BeautifulSoup(html_content, 'html.parser')
    rows = soup.select('tr.stockTrMobile')

    stock_data = []

    for row in rows:
        p_tags = row.select('td p.stockInfoMobile')
        if len(p_tags) >= 2:
            stock_name = p_tags[0].text.strip()
            stock_ticker = p_tags[1].text.strip().replace('(', '').replace(')', '')
            
            # 종목명과 티커를 함께 저장하거나, 종목명만 저장할 수 있습니다.
            # 요청하신 대로 '파일명'을 컬럼명으로 하기 위해 종목명을 리스트에 담습니다.
            stock_data.append(stock_name)

    # DataFrame 생성
    df = pd.DataFrame({file_name: stock_data})
    
    # 중복 제거
    df = df.drop_duplicates()
    
    return df

if __name__ == "__main__":
    from collections import defaultdict
    import glob

    # 타겟 파일 리스트 (glob 사용)
    target_files = glob.glob(r'data/theme_html/*.html')
    
    # 종목별 테마 리스트 저장
    stock_themes = defaultdict(list)

    print(f"Checking {len(target_files)} files...")

    # 데이터 수집 (테마 -> 종목 리스트, 종목 -> 테마 리스트)
    theme_stocks_map = {}
    
    for file_path in target_files:
        df = extract_stocks_from_html(file_path)
        if df is not None:
            theme_name = df.columns[0]
            stocks = df[theme_name].tolist()
            
            # 테마별 종목 리스트 저장
            theme_stocks_map[theme_name] = stocks
            
            for stock in stocks:
                stock_themes[stock].append(theme_name)
    
    # 2개 이상의 테마에 포함된 종목 필터링 (중복 종목)
    common_stocks = {k: v for k, v in stock_themes.items() if len(v) >= 2}
    common_stock_names = set(common_stocks.keys())
    
    print(f"\nTotal Unique Stocks: {len(stock_themes)}")
    print(f"Common Stocks (appearing in > 1 themes): {len(common_stocks)}")

    # 콘솔 출력
    for stock in sorted(common_stocks.keys()):
        themes = common_stocks[stock]
        print(f"{stock}: {themes}")

    # DataFrame 생성 (키를 컬럼으로 변환하기 위해 orient='index' 후 Transpose)
    # 1. 테마상세 (컬럼: 테마명)
    df_detail = pd.DataFrame.from_dict(theme_stocks_map, orient='index').T
    
    # 2. 중복종목 (행: 종목명, A열이 종목명이 되도록 Transpose 제거)
    df_common = pd.DataFrame.from_dict(common_stocks, orient='index')
    # 인덱스를 컬럼으로 빼내고 컬럼명 지정
    df_common.reset_index(inplace=True)
    theme_cols = [f'테마{i+1}' for i in range(df_common.shape[1]-1)]
    df_common.columns = ['종목명'] + theme_cols

    print(f"\nDataFrame Shapes - Detail: {df_detail.shape}, Common: {df_common.shape}")

    # 엑셀 파일로 저장 (스타일 없이 데이터만 먼저 저장)
    output_file = 'theme_stocks.xlsx'
    
    import sys
    
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
        except PermissionError:
            print(f"Error: {output_file} is open. Please close it.")
            sys.exit(1)

    try:
        # 1차 저장: 데이터만 저장
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_detail.to_excel(writer, sheet_name='테마상세', index=False)
            df_common.to_excel(writer, sheet_name='중복종목', index=False)
            
        # 2차 작업: openpyxl로 열어서 스타일 적용 및 컬럼 너비 조정
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(output_file)
        ws_detail = wb['테마상세']
        ws_common = wb['중복종목']
        
        # 빨간 배경 스타일 정의
        red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
        
        # 테마상세 시트 순회하며 스타일 적용
        # 헤더(1행) 제외하고 2행부터 시작
        for row in ws_detail.iter_rows(min_row=2):
            for cell in row:
                if cell.value and cell.value in common_stock_names:
                    cell.fill = red_fill
        
        # 컬럼 너비 자동 조정 함수
        def adjust_auto_width(ws):
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                # 한글 등 멀티바이트 문자 고려하여 여유 있게 설정
                adjusted_width = (max_length + 2) * 1.5
                ws.column_dimensions[column].width = adjusted_width

        # 두 시트에 모두 적용
        adjust_auto_width(ws_detail)
        adjust_auto_width(ws_common)
                    
        wb.save(output_file)
        print(f"\n엑셀 파일 저장, 스타일 및 너비 조정 완료: {output_file}")
            
    except Exception as e:
        print(f"\n엑셀 저장 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
